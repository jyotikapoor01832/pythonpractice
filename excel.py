from openpyxl import Workbook

wb=Workbook()
ws=wb.active

ws["A1"]="HELLO"
ws["B2"]="WORLD"
ws["C3"]="FROM"
ws["D4"]="PYTHON"
wb.save('sample2.xlsx')